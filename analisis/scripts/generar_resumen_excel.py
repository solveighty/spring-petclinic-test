"""
Generar resumen de estadística descriptiva en Excel
Media, Mediana, Desviación Estándar por métrica y grupo
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# 1. CARGAR DATOS
# ============================================================================

BASE_PATH = Path("./unit_tests_metrics")
FUNCTIONAL_PATH = Path("./functional_tests_metrics")

print("Cargando datos...")

dfs = []

unitarias_files = [
    "IA_OwnerAddPetDiffblueTest.csv",
    "IA_OwnerGetPetDiffblueTest.csv",
    "IA_PetValidatorDiffblueTest.csv",
    "Manual_OwnerAddPetUnitManualTest.csv",
    "Manual_OwnerGetPetUnitManualTest.csv",
    "Manual_PetValidatorUnitManualTest.csv",
]

funcionales_files = [
    "IA_OwnerControllerShowOwnerTestIA.csv",
    "IA_PetControllerProcessCreationFormTestIA.csv",
    "IA_VisitControllerProcessNewVisitFormTestIA.csv",
    "Manual_ShowOwnerManualTest.csv",
    "Manual_ProcessCreationFormManualTest.csv",
    "Manual_ProcessNewVisitFormManualTest.csv",
]

for file in unitarias_files:
    filepath = BASE_PATH / file
    if filepath.exists():
        df_temp = pd.read_csv(filepath)
        df_temp['category'] = 'Unitarias'
        dfs.append(df_temp)

for file in funcionales_files:
    filepath = FUNCTIONAL_PATH / file
    if filepath.exists():
        df_temp = pd.read_csv(filepath)
        df_temp['category'] = 'Funcionales'
        dfs.append(df_temp)

df = pd.concat(dfs, ignore_index=True)
print(f"Total: {len(df)} registros cargados")

metricas = ['instr_pct', 'branch_pct', 'mutation_score', 'time_seconds']

# ============================================================================
# 1. CREAR RESUMEN GLOBAL
# ============================================================================

resumen_global = []

for metrica in metricas:
    for grupo in ['Manual', 'IA']:
        datos = df[df['group'] == grupo][metrica]
        
        resumen_global.append({
            'Metrica': metrica,
            'Grupo': grupo,
            'n': len(datos),
            'Media': datos.mean(),
            'Mediana': datos.median(),
            'Desv.Est': datos.std(),
            'Min': datos.min(),
            'Max': datos.max(),
            'Q1': datos.quantile(0.25),
            'Q3': datos.quantile(0.75),
            'IQR': datos.quantile(0.75) - datos.quantile(0.25),
        })

df_resumen_global = pd.DataFrame(resumen_global)

# ============================================================================
# 2. CREAR RESUMEN POR CATEGORÍA
# ============================================================================

resumen_categorias = []

for categoria in ['Unitarias', 'Funcionales']:
    df_cat = df[df['category'] == categoria]
    
    for metrica in metricas:
        for grupo in ['Manual', 'IA']:
            datos = df_cat[df_cat['group'] == grupo][metrica]
            
            if len(datos) > 0:
                resumen_categorias.append({
                    'Categoria': categoria,
                    'Metrica': metrica,
                    'Grupo': grupo,
                    'n': len(datos),
                    'Media': datos.mean(),
                    'Mediana': datos.median(),
                    'Desv.Est': datos.std(),
                    'Min': datos.min(),
                    'Max': datos.max(),
                    'Q1': datos.quantile(0.25),
                    'Q3': datos.quantile(0.75),
                    'IQR': datos.quantile(0.75) - datos.quantile(0.25),
                })

df_resumen_categorias = pd.DataFrame(resumen_categorias)

# ============================================================================
# 3. CREAR RESUMEN POR CLASE DE TEST
# ============================================================================

resumen_clases = []

test_classes = sorted(df['test_class'].unique())

for test_class in test_classes:
    df_test = df[df['test_class'] == test_class]
    
    for metrica in metricas:
        for grupo in ['Manual', 'IA']:
            datos = df_test[df_test['group'] == grupo][metrica]
            
            if len(datos) > 0:
                resumen_clases.append({
                    'Clase': test_class,
                    'Metrica': metrica,
                    'Grupo': grupo,
                    'n': len(datos),
                    'Media': datos.mean(),
                    'Mediana': datos.median(),
                    'Desv.Est': datos.std(),
                    'Min': datos.min(),
                    'Max': datos.max(),
                    'Q1': datos.quantile(0.25),
                    'Q3': datos.quantile(0.75),
                    'IQR': datos.quantile(0.75) - datos.quantile(0.25),
                })

df_resumen_clases = pd.DataFrame(resumen_clases)

# ============================================================================
# 4. CREAR COMPARATIVA MANUAL vs IA
# ============================================================================

comparativa = []

for metrica in metricas:
    manual = df[df['group'] == 'Manual'][metrica]
    ia = df[df['group'] == 'IA'][metrica]
    
    diff_media = ia.mean() - manual.mean()
    diff_media_pct = (diff_media / manual.mean() * 100) if manual.mean() != 0 else 0
    
    diff_mediana = ia.median() - manual.median()
    diff_mediana_pct = (diff_mediana / manual.median() * 100) if manual.median() != 0 else 0
    
    comparativa.append({
        'Metrica': metrica,
        'Manual_Media': manual.mean(),
        'IA_Media': ia.mean(),
        'Diferencia_Media': diff_media,
        'Diferencia_Media_%': diff_media_pct,
        'Manual_Mediana': manual.median(),
        'IA_Mediana': ia.median(),
        'Diferencia_Mediana': diff_mediana,
        'Diferencia_Mediana_%': diff_mediana_pct,
        'Manual_StdDev': manual.std(),
        'IA_StdDev': ia.std(),
    })

df_comparativa = pd.DataFrame(comparativa)

# ============================================================================
# 5. ESCRIBIR EN EXCEL (agregar a archivo existente)
# ============================================================================

print("Creando Excel con resumen estadístico...")

try:
    # Intentar cargar archivo existente
    with pd.ExcelWriter('ESTADISTICA_DESCRIPTIVA.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_resumen_global.round(4).to_excel(writer, sheet_name='Resumen Global', index=False)
        df_resumen_categorias.round(4).to_excel(writer, sheet_name='Por Categoria', index=False)
        df_resumen_clases.round(4).to_excel(writer, sheet_name='Por Clase', index=False)
        df_comparativa.round(4).to_excel(writer, sheet_name='Comparativa Manual vs IA', index=False)
    
    print("OK: Archivos agregados a ESTADISTICA_DESCRIPTIVA.xlsx")
    
except Exception as e:
    print(f"Error: {e}")
    print("Creando archivo nuevo...")
    
    with pd.ExcelWriter('ESTADISTICA_DESCRIPTIVA.xlsx', engine='openpyxl') as writer:
        df_resumen_global.round(4).to_excel(writer, sheet_name='Resumen Global', index=False)
        df_resumen_categorias.round(4).to_excel(writer, sheet_name='Por Categoria', index=False)
        df_resumen_clases.round(4).to_excel(writer, sheet_name='Por Clase', index=False)
        df_comparativa.round(4).to_excel(writer, sheet_name='Comparativa Manual vs IA', index=False)
        df.to_excel(writer, sheet_name='Datos Crudos', index=False)
    
    print("OK: Archivo ESTADISTICA_DESCRIPTIVA.xlsx creado")

# ============================================================================
# 6. APLICAR FORMATO A LAS HOJAS
# ============================================================================

print("Aplicando formato...")

wb = load_workbook('ESTADISTICA_DESCRIPTIVA.xlsx')

# Estilos
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for sheet_name in wb.sheetnames:
    if sheet_name != 'Datos Crudos':
        ws = wb[sheet_name]
        
        # Aplicar formato a headers
        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
            for cell in col:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Aplicar bordes a todas las celdas
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Formatear números
                if isinstance(cell.value, float):
                    cell.number_format = '0.0000'

wb.save('ESTADISTICA_DESCRIPTIVA.xlsx')
print("OK: Formato aplicado")

# ============================================================================
# 7. MOSTRAR RESUMEN EN CONSOLA
# ============================================================================

print("\n" + "="*100)
print("RESUMEN GLOBAL - MANUAL vs IA")
print("="*100)
print(df_comparativa.to_string(index=False))

print("\n" + "="*100)
print("ARCHIVOS GENERADOS/ACTUALIZADOS")
print("="*100)
print("1. ESTADISTICA_DESCRIPTIVA.xlsx")
print("   - Hoja 'Resumen Global': Todas las métricas, ambos grupos")
print("   - Hoja 'Por Categoria': Unitarias y Funcionales")
print("   - Hoja 'Por Clase': Cada clase de test (OwnerAddPet, OwnerGetPet, etc.)")
print("   - Hoja 'Comparativa Manual vs IA': Diferencias y porcentajes")
print("   - Hoja 'Datos Crudos': Todos los 2,480 registros")
print("\nCompleto.")
