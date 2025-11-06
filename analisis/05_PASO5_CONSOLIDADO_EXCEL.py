#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PASO 5: CONSOLIDADO EXCEL PARA CAPÍTULO 4
===========================================

Genera un único archivo Excel con TODAS las tablas necesarias para la tesis:
- Tabla 4.1: Descriptivos N=2,480
- Tabla 4.2: Descriptivos N=12
- Tabla 4.3: Shapiro-Wilk (N=12)
- Tabla 4.4: Levene (N=12)
- Tabla 4.5: t-Student/Welch (N=12)
- Tabla 4.6: Cohen's d (N=12)
- Tabla 4.7: Resumen de Supuestos

Formato: Listo para copiar/pegar a Word/Google Docs
"""

import pandas as pd
import numpy as np
from scipy.stats import shapiro, levene, ttest_ind
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

print("=" * 100)
print("PASO 5: CONSOLIDADO EXCEL PARA CAPÍTULO 4")
print("=" * 100)

# ─────────────────────────────────────────────────────────────────────────────
# CARGAR DATOS
# ─────────────────────────────────────────────────────────────────────────────

print("\n[PASO 0] Cargando datos...")

df = pd.read_csv('datos_consolidados.csv')
print(f"  ✓ Datos: {len(df)} registros")

METRICAS = ['instr_pct', 'branch_pct', 'mutation_score', 'time_seconds']

# N=12
df_promedios_raw = df.groupby('test_name')[METRICAS].mean().reset_index()
df_group = df.groupby('test_name')['group'].first().reset_index()
df_promedios = df_promedios_raw.merge(df_group, on='test_name')

datos_n12 = {
    'Manual': df_promedios[df_promedios['group'] == 'Manual'],
    'IA': df_promedios[df_promedios['group'] == 'IA']
}

print(f"  ✓ Promedios: {len(df_promedios)} tests")

# ─────────────────────────────────────────────────────────────────────────────
# CREAR DATAFRAMES PARA CADA TABLA
# ─────────────────────────────────────────────────────────────────────────────

print("\n[PASO 1] Preparando tablas...")

# ─────────────────────────────────────────────────────────────────────────────
# TABLA 4.1: DESCRIPTIVOS N=2,480
# ─────────────────────────────────────────────────────────────────────────────

tabla_descriptivos_2480 = []

for metrica in METRICAS:
    datos_manual = df[df['group'] == 'Manual'][metrica]
    datos_ia = df[df['group'] == 'IA'][metrica]
    
    tabla_descriptivos_2480.append({
        'Métrica': metrica,
        'Grupo': 'Manual',
        'N': len(datos_manual),
        'Media': round(datos_manual.mean(), 4),
        'Mediana': round(datos_manual.median(), 4),
        'Desv. Est.': round(datos_manual.std(), 4),
        'Mín': round(datos_manual.min(), 4),
        'Máx': round(datos_manual.max(), 4),
        'Q1': round(datos_manual.quantile(0.25), 4),
        'Q3': round(datos_manual.quantile(0.75), 4)
    })
    
    tabla_descriptivos_2480.append({
        'Métrica': metrica,
        'Grupo': 'AI',
        'N': len(datos_ia),
        'Media': round(datos_ia.mean(), 4),
        'Mediana': round(datos_ia.median(), 4),
        'Desv. Est.': round(datos_ia.std(), 4),
        'Mín': round(datos_ia.min(), 4),
        'Máx': round(datos_ia.max(), 4),
        'Q1': round(datos_ia.quantile(0.25), 4),
        'Q3': round(datos_ia.quantile(0.75), 4)
    })

df_tabla_41 = pd.DataFrame(tabla_descriptivos_2480)

# ─────────────────────────────────────────────────────────────────────────────
# TABLA 4.2: DESCRIPTIVOS N=12
# ─────────────────────────────────────────────────────────────────────────────

tabla_descriptivos_12 = []

for metrica in METRICAS:
    datos_manual = datos_n12['Manual'][metrica]
    datos_ia = datos_n12['IA'][metrica]
    
    tabla_descriptivos_12.append({
        'Métrica': metrica,
        'Grupo': 'Manual',
        'N': len(datos_manual),
        'Media': round(datos_manual.mean(), 4),
        'Mediana': round(datos_manual.median(), 4),
        'Desv. Est.': round(datos_manual.std(), 4),
        'Mín': round(datos_manual.min(), 4),
        'Máx': round(datos_manual.max(), 4)
    })
    
    tabla_descriptivos_12.append({
        'Métrica': metrica,
        'Grupo': 'AI',
        'N': len(datos_ia),
        'Media': round(datos_ia.mean(), 4),
        'Mediana': round(datos_ia.median(), 4),
        'Desv. Est.': round(datos_ia.std(), 4),
        'Mín': round(datos_ia.min(), 4),
        'Máx': round(datos_ia.max(), 4)
    })

df_tabla_42 = pd.DataFrame(tabla_descriptivos_12)

# ─────────────────────────────────────────────────────────────────────────────
# TABLA 4.3: SHAPIRO-WILK (N=12)
# ─────────────────────────────────────────────────────────────────────────────

tabla_shapiro = []

for metrica in METRICAS:
    datos_manual = datos_n12['Manual'][metrica].values
    datos_ia = datos_n12['IA'][metrica].values
    
    w_m, p_m = shapiro(datos_manual)
    w_i, p_i = shapiro(datos_ia)
    
    tabla_shapiro.append({
        'Métrica': metrica,
        'Grupo': 'Manual',
        'W-statistic': round(w_m, 6),
        'p-value': round(p_m, 6),
        'Interpretación': 'Normal (p > 0.05)' if p_m > 0.05 else 'NO Normal (p < 0.05)',
        'Decisión': 'Rechazar H0' if p_m < 0.05 else 'Aceptar H0'
    })
    
    tabla_shapiro.append({
        'Métrica': metrica,
        'Grupo': 'AI',
        'W-statistic': round(w_i, 6),
        'p-value': round(p_i, 6),
        'Interpretación': 'Normal (p > 0.05)' if p_i > 0.05 else 'NO Normal (p < 0.05)',
        'Decisión': 'Rechazar H0' if p_i < 0.05 else 'Aceptar H0'
    })

df_tabla_43 = pd.DataFrame(tabla_shapiro)

# ─────────────────────────────────────────────────────────────────────────────
# TABLA 4.4: LEVENE (N=12)
# ─────────────────────────────────────────────────────────────────────────────

tabla_levene = []

for metrica in METRICAS:
    datos_manual = datos_n12['Manual'][metrica]
    datos_ia = datos_n12['IA'][metrica]
    
    stat, pval = levene(datos_manual, datos_ia)
    
    tabla_levene.append({
        'Métrica': metrica,
        'Test Statistic': round(stat, 6),
        'p-value': round(pval, 6),
        'Varianzas': 'Iguales (p ≥ 0.05)' if pval >= 0.05 else 'Desiguales (p < 0.05)',
        'Test a usar': 't-Student estándar' if pval >= 0.05 else 't-Student Welch'
    })

df_tabla_44 = pd.DataFrame(tabla_levene)

# ─────────────────────────────────────────────────────────────────────────────
# TABLA 4.5: t-STUDENT/WELCH (N=12)
# ─────────────────────────────────────────────────────────────────────────────

# Cargar desde Excel que ya tiene los cálculos
excel_paso3 = pd.read_excel('03_PASO3_HIPOTESIS_T_STUDENT.xlsx', sheet_name=0)

tabla_ttest = []

for _, row in excel_paso3.iterrows():
    tabla_ttest.append({
        'Métrica': row['metrica'],
        'N Manual': int(row['n_manual']),
        'N AI': int(row['n_ia']),
        'Media Manual': round(row['media_manual'], 4),
        'Media AI': round(row['media_ia'], 4),
        'Diferencia (%)': round(row['pct_diferencia'], 2),
        't-statistic': round(row['t_statistic'], 4),
        'p-value': round(row['p_value'], 6),
        'Significativo (α=0.05)': 'No' if row['p_value'] >= 0.05 else 'Sí',
        'Test usado': row['test_usado']
    })

df_tabla_45 = pd.DataFrame(tabla_ttest)

# ─────────────────────────────────────────────────────────────────────────────
# TABLA 4.6: COHEN'S D (N=12)
# ─────────────────────────────────────────────────────────────────────────────

tabla_cohens_d = []

for _, row in excel_paso3.iterrows():
    d = row['cohens_d']
    abs_d = abs(d)
    
    if abs_d < 0.2:
        magnitud = 'Negligible'
    elif abs_d < 0.5:
        magnitud = 'Small'
    elif abs_d < 0.8:
        magnitud = 'Medium'
    else:
        magnitud = 'Large'
    
    tabla_cohens_d.append({
        'Métrica': row['metrica'],
        "Cohen's d": round(d, 4),
        'Magnitud': magnitud,
        'Dirección': 'Manual > AI' if d > 0 else 'Manual < AI',
        'Interpretación': row['interpretacion_d']
    })

df_tabla_46 = pd.DataFrame(tabla_cohens_d)

# ─────────────────────────────────────────────────────────────────────────────
# TABLA 4.7: RESUMEN DE SUPUESTOS
# ─────────────────────────────────────────────────────────────────────────────

tabla_supuestos = []

for metrica in METRICAS:
    # Shapiro
    datos_manual = datos_n12['Manual'][metrica].values
    datos_ia = datos_n12['IA'][metrica].values
    w_m, p_shapiro_m = shapiro(datos_manual)
    w_i, p_shapiro_i = shapiro(datos_ia)
    
    # Levene
    stat, p_levene = levene(datos_manual, datos_ia)
    
    # Decisión
    shapiro_ok = (p_shapiro_m > 0.05) and (p_shapiro_i > 0.05)
    levene_ok = p_levene >= 0.05
    
    tabla_supuestos.append({
        'Métrica': metrica,
        'Normalidad (Shapiro-Wilk)': '✓ Sí' if shapiro_ok else '✗ No',
        'Igualdad de Varianzas (Levene)': '✓ Sí' if levene_ok else '✗ No',
        'Supuestos Cumplidos': '✓ Sí' if (shapiro_ok and levene_ok) else '✗ Parcialmente',
        'Test Estadístico': 't-Student estándar' if (shapiro_ok and levene_ok) else 
                           ('t-Student Welch' if shapiro_ok else 'Mann-Whitney U')
    })

df_tabla_47 = pd.DataFrame(tabla_supuestos)

# ─────────────────────────────────────────────────────────────────────────────
# CREAR EXCEL CON MÚLTIPLES HOJAS
# ─────────────────────────────────────────────────────────────────────────────

print("\n[PASO 2] Creando archivo Excel...")

with pd.ExcelWriter('05_PASO5_CONSOLIDADO_CAPITULO4.xlsx', engine='openpyxl') as writer:
    
    # Tabla 4.1
    df_tabla_41.to_excel(writer, sheet_name='Tabla 4.1 - Desc 2480', index=False)
    print("  ✓ Tabla 4.1: Descriptivos N=2,480")
    
    # Tabla 4.2
    df_tabla_42.to_excel(writer, sheet_name='Tabla 4.2 - Desc 12', index=False)
    print("  ✓ Tabla 4.2: Descriptivos N=12")
    
    # Tabla 4.3
    df_tabla_43.to_excel(writer, sheet_name='Tabla 4.3 - Shapiro', index=False)
    print("  ✓ Tabla 4.3: Shapiro-Wilk")
    
    # Tabla 4.4
    df_tabla_44.to_excel(writer, sheet_name='Tabla 4.4 - Levene', index=False)
    print("  ✓ Tabla 4.4: Levene")
    
    # Tabla 4.5
    df_tabla_45.to_excel(writer, sheet_name='Tabla 4.5 - t-Student', index=False)
    print("  ✓ Tabla 4.5: t-Student/Welch")
    
    # Tabla 4.6
    df_tabla_46.to_excel(writer, sheet_name='Tabla 4.6 - Cohen\'s d', index=False)
    print("  ✓ Tabla 4.6: Cohen's d")
    
    # Tabla 4.7
    df_tabla_47.to_excel(writer, sheet_name='Tabla 4.7 - Supuestos', index=False)
    print("  ✓ Tabla 4.7: Resumen de Supuestos")

# ─────────────────────────────────────────────────────────────────────────────
# FORMATEAR EXCEL
# ─────────────────────────────────────────────────────────────────────────────

print("\n[PASO 3] Formateando Excel...")

from openpyxl import load_workbook

wb = load_workbook('05_PASO5_CONSOLIDADO_CAPITULO4.xlsx')

# Estilos
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Formato de encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Formato de datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Formatear números
            if isinstance(cell.value, float):
                if cell.value < 1:
                    cell.number_format = '0.0000'
                else:
                    cell.number_format = '0.00'

wb.save('05_PASO5_CONSOLIDADO_CAPITULO4.xlsx')
print("  ✓ Formato aplicado")

# ─────────────────────────────────────────────────────────────────────────────
# RESUMEN FINAL
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "=" * 100)
print("✓ PASO 5 COMPLETADO: EXCEL CONSOLIDADO CREADO")
print("=" * 100)

print("""
ARCHIVO GENERADO:
  📊 05_PASO5_CONSOLIDADO_CAPITULO4.xlsx

CONTENIDO (7 hojas):
  ✓ Tabla 4.1: Descriptivos N=2,480 (Media, Mediana, Desv. Est., Mín, Máx, Q1, Q3)
  ✓ Tabla 4.2: Descriptivos N=12 (Media, Mediana, Desv. Est., Mín, Máx)
  ✓ Tabla 4.3: Shapiro-Wilk (W-stat, p-value, Interpretación)
  ✓ Tabla 4.4: Levene (Test Stat, p-value, Decisión)
  ✓ Tabla 4.5: t-Student/Welch (t-stat, p-value, Diferencia %, Test usado)
  ✓ Tabla 4.6: Cohen's d (Magnitud, Dirección, Interpretación)
  ✓ Tabla 4.7: Resumen de Supuestos (Normalidad, Varianzas, Decisión)

FORMATO:
  ✓ Encabezados: Azul con texto blanco, bold
  ✓ Bordes: Todas las celdas delimitadas
  ✓ Ancho automático: Ajustado al contenido
  ✓ Números: 4 decimales para p-values, 2 para porcentajes
  ✓ Alineación: Centrada en todas las celdas

CÓMO USAR:
  1. Abrir 05_PASO5_CONSOLIDADO_CAPITULO4.xlsx
  2. Seleccionar la tabla deseada (ej: "Tabla 4.1 - Desc 2480")
  3. Copiar tabla (Ctrl+A en la hoja, luego Ctrl+C)
  4. Pegar en Word/Google Docs (Ctrl+V)
  5. Tabla se formatea automáticamente

PRÓXIMO PASO:
  - PASO 6: Redactar Capítulo 4 con interpretaciones

════════════════════════════════════════════════════════════════════════════════
""")

print("=" * 100)
