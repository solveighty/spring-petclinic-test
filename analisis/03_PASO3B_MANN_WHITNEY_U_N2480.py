#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
════════════════════════════════════════════════════════════════════════════════
PASO 3B: PRUEBA DE MANN-WHITNEY U (DATOS BRUTOS N=2,480)
════════════════════════════════════════════════════════════════════════════════

Objetivo:
  Validar con test no-paramétrico (Mann-Whitney U) que los resultados obtenidos
  con t-Student/Welch en datos agregados (N=12) son robustos y concordantes
  con el análisis de datos brutos (N=2,480).

Justificación:
  • N=2,480 RECHAZA normalidad (Shapiro-Wilk p < 0.001)
  • Se requiere test no-paramétrico para validación
  • Mann-Whitney U no asume normalidad
  • Compara medianas en lugar de medias

Métrica de comparación:
  Si p-values de Mann-Whitney U ≈ p-values de t-Student → Conclusiones ROBUSTAS

════════════════════════════════════════════════════════════════════════════════
"""

import pandas as pd
import numpy as np
from scipy.stats import mannwhitneyu
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

print("════════════════════════════════════════════════════════════════════════════════")
print("PASO 3B: PRUEBA DE MANN-WHITNEY U (DATOS BRUTOS N=2,480)")
print("════════════════════════════════════════════════════════════════════════════════\n")

# ════════════════════════════════════════════════════════════════════════════════
# PASO 0: CARGAR DATOS
# ════════════════════════════════════════════════════════════════════════════════

print("[PASO 0] Cargando datos brutos...")
df = pd.read_csv('datos_consolidados.csv')
print(f"  ✓ Registros cargados: {len(df):,}")
print(f"  ✓ Grupos: {df['group'].unique().tolist()}")
print(f"  ✓ Métricas disponibles: instr_pct, branch_pct, mutation_score, time_seconds\n")

# ════════════════════════════════════════════════════════════════════════════════
# PASO 1: EXTRAER DATOS POR GRUPO Y MÉTRICA
# ════════════════════════════════════════════════════════════════════════════════

print("[PASO 1] Preparando datos para Mann-Whitney U...")

# Separar por grupo
manual_data = df[df['group'] == 'Manual']
ia_data = df[df['group'] == 'IA']

print(f"  ✓ Registros Manual: {len(manual_data):,}")
print(f"  ✓ Registros AI: {len(ia_data):,}\n")

# ════════════════════════════════════════════════════════════════════════════════
# PASO 2: CALCULAR MANN-WHITNEY U PARA CADA MÉTRICA
# ════════════════════════════════════════════════════════════════════════════════

print("[PASO 2] Ejecutando Mann-Whitney U para cada métrica...")

metricas = ['instr_pct', 'branch_pct', 'mutation_score', 'time_seconds']
labels_metricas = [
    'Instruction Coverage (%)',
    'Branch Coverage (%)',
    'Mutation Score (%)',
    'Time (seconds)'
]

resultados_mw = []

for metrica, label in zip(metricas, labels_metricas):
    
    # Extraer datos
    manual_vals = manual_data[metrica].dropna()
    ia_vals = ia_data[metrica].dropna()
    
    # Calcular estadísticos descriptivos
    median_manual = manual_vals.median()
    median_ia = ia_vals.median()
    mean_manual = manual_vals.mean()
    mean_ia = ia_vals.mean()
    std_manual = manual_vals.std()
    std_ia = ia_vals.std()
    
    # Mann-Whitney U test (two-sided)
    u_stat, p_value = mannwhitneyu(manual_vals, ia_vals, alternative='two-sided')
    
    # Calcular tamaño del efecto (r = Z / sqrt(N))
    from scipy.stats import norm
    # Aproximación: Z = (U - E[U]) / sqrt(Var[U])
    n1, n2 = len(manual_vals), len(ia_vals)
    E_U = (n1 * n2) / 2
    Var_U = (n1 * n2 * (n1 + n2 + 1)) / 12
    Z = (u_stat - E_U) / np.sqrt(Var_U)
    r_effect = abs(Z) / np.sqrt(n1 + n2)  # Tamaño del efecto (r)
    
    # Interpretación de p-value
    es_significativo = "Sí" if p_value < 0.05 else "No"
    
    # Almacenar resultados
    resultados_mw.append({
        'Métrica': label,
        'N Manual': n1,
        'N AI': n2,
        'Mediana Manual': median_manual,
        'Mediana AI': median_ia,
        'Media Manual': mean_manual,
        'Media AI': mean_ia,
        'Desv.Est. Manual': std_manual,
        'Desv.Est. AI': std_ia,
        'U-statistic': u_stat,
        'Z-score': Z,
        'p-value': p_value,
        'r (effect size)': r_effect,
        'Significativo (α=0.05)': es_significativo,
    })
    
    print(f"\n  {label}:")
    print(f"    N Manual: {n1:,}  |  N AI: {n2:,}")
    print(f"    Mediana Manual: {median_manual:.4f}  |  Mediana AI: {median_ia:.4f}")
    print(f"    U-statistic: {u_stat:.2f}")
    print(f"    p-value: {p_value:.6f}")
    print(f"    Significativo: {es_significativo} (α=0.05)")

# ════════════════════════════════════════════════════════════════════════════════
# PASO 3: CREAR DATAFRAME DE RESULTADOS
# ════════════════════════════════════════════════════════════════════════════════

print("\n[PASO 3] Creando tabla de resultados...\n")
df_resultados = pd.DataFrame(resultados_mw)
print(df_resultados.to_string())

# ════════════════════════════════════════════════════════════════════════════════
# PASO 4: CREAR ARCHIVO EXCEL CON FORMATOS
# ════════════════════════════════════════════════════════════════════════════════

print("\n[PASO 4] Generando archivo Excel...\n")

wb = Workbook()

# Hoja 1: Resultados Mann-Whitney U
ws1 = wb.active
ws1.title = "Mann-Whitney_U_N2480"

# Estilos
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Escribir headers
headers = [
    'Métrica',
    'N Manual',
    'N AI',
    'Mediana Manual',
    'Mediana AI',
    'Media Manual',
    'Media AI',
    'Desv.Est. Manual',
    'Desv.Est. AI',
    'U-statistic',
    'Z-score',
    'p-value',
    'r (effect size)',
    'Significativo'
]

for col_idx, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = border

# Escribir datos
for row_idx, row_data in enumerate(resultados_mw, 2):
    ws1.cell(row=row_idx, column=1).value = row_data['Métrica']
    ws1.cell(row=row_idx, column=2).value = row_data['N Manual']
    ws1.cell(row=row_idx, column=3).value = row_data['N AI']
    ws1.cell(row=row_idx, column=4).value = round(row_data['Mediana Manual'], 4)
    ws1.cell(row=row_idx, column=5).value = round(row_data['Mediana AI'], 4)
    ws1.cell(row=row_idx, column=6).value = round(row_data['Media Manual'], 4)
    ws1.cell(row=row_idx, column=7).value = round(row_data['Media AI'], 4)
    ws1.cell(row=row_idx, column=8).value = round(row_data['Desv.Est. Manual'], 4)
    ws1.cell(row=row_idx, column=9).value = round(row_data['Desv.Est. AI'], 4)
    ws1.cell(row=row_idx, column=10).value = round(row_data['U-statistic'], 2)
    ws1.cell(row=row_idx, column=11).value = round(row_data['Z-score'], 4)
    ws1.cell(row=row_idx, column=12).value = row_data['p-value']
    ws1.cell(row=row_idx, column=13).value = round(row_data['r (effect size)'], 4)
    ws1.cell(row=row_idx, column=14).value = row_data['Significativo (α=0.05)']
    
    # Aplicar estilos a datos
    for col_idx in range(1, len(headers) + 1):
        cell = ws1.cell(row=row_idx, column=col_idx)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Formato para p-value (científico)
        if col_idx == 12:
            cell.number_format = '0.0000E+00'
        elif col_idx in [4, 5, 6, 7, 8, 9, 10, 11, 13]:
            cell.number_format = '0.0000'

# Ajustar ancho de columnas
for col_idx in range(1, len(headers) + 1):
    ws1.column_dimensions[get_column_letter(col_idx)].width = 15

# Hoja 2: Interpretación
ws2 = wb.create_sheet("Interpretacion")

interpretacion_texto = """
MANN-WHITNEY U TEST - INTERPRETACIÓN

════════════════════════════════════════════════════════════════════════════════

MÉTODO ESTADÍSTICO:
  • Prueba no-paramétrica (no asume normalidad)
  • Compara medianas en lugar de medias
  • Válida para datos que violan normalidad
  • Robusto a outliers

HIPÓTESIS:
  H₀: La distribución de Manual = La distribución de AI (no hay diferencia)
  H₁: La distribución de Manual ≠ La distribución de AI (hay diferencia)

NIVEL DE SIGNIFICANCIA:
  α = 0.05 (5% de error permitido)

INTERPRETACIÓN DE RESULTADOS:
  • Si p-value > 0.05 → NO rechazamos H₀ → NO hay diferencia significativa
  • Si p-value < 0.05 → Rechazamos H₀ → Hay diferencia significativa

TAMAÑO DEL EFECTO (r):
  • r < 0.1   → Efecto negligible
  • 0.1 ≤ r < 0.3  → Efecto pequeño
  • 0.3 ≤ r < 0.5  → Efecto mediano
  • r ≥ 0.5   → Efecto grande

════════════════════════════════════════════════════════════════════════════════

VALIDACIÓN DE CONCORDANCIA:

Si los p-values de Mann-Whitney U son SIMILARES a los de t-Student:
  ✓ Las conclusiones son ROBUSTAS
  ✓ No hay diferencias por violación de normalidad
  ✓ Conclusiones son válidas independientemente del método

Si hay GRANDES DIFERENCIAS entre p-values:
  ✓ Requiere investigación adicional
  ✓ Pueden existir outliers significativos
  ✓ Los datos brutos pueden necesitar transformación

════════════════════════════════════════════════════════════════════════════════
"""

ws2.column_dimensions['A'].width = 100
cell = ws2['A1']
cell.value = interpretacion_texto
cell.alignment = Alignment(wrap_text=True, vertical="top")

# Hoja 3: Comparación con t-Student
ws3 = wb.create_sheet("Comparacion_t-Student_vs_MW")

# Cargar resultados de t-Student
df_tstudent = pd.read_excel('03_PASO3_HIPOTESIS_T_STUDENT.xlsx', sheet_name=0)

# Crear tabla comparativa
headers_comp = [
    'Métrica',
    't-Student p-value',
    'Mann-Whitney p-value',
    'Diferencia',
    'Concordancia',
    'Conclusión'
]

for col_idx, header in enumerate(headers_comp, 1):
    cell = ws3.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = border

# Llenar datos comparativos
for row_idx, resultado_mw in enumerate(resultados_mw, 2):
    # Obtener índice de métrica
    metrica_row = row_idx - 2
    metrica_ts = metricas[metrica_row]
    p_tstudent = df_tstudent.iloc[metrica_row]['p_value']
    p_mw = resultado_mw['p-value']
    
    # Calcular diferencia
    diff = abs(p_tstudent - p_mw)
    
    # Evaluar concordancia
    if (p_tstudent > 0.05 and p_mw > 0.05) or (p_tstudent < 0.05 and p_mw < 0.05):
        concordancia = "✓ SÍ"
        conclusion = "Conclusiones idénticas"
    else:
        concordancia = "⚠ PARCIAL"
        conclusion = "Requiere análisis"
    
    ws3.cell(row=row_idx, column=1).value = resultado_mw['Métrica']
    ws3.cell(row=row_idx, column=2).value = p_tstudent
    ws3.cell(row=row_idx, column=3).value = p_mw
    ws3.cell(row=row_idx, column=4).value = diff
    ws3.cell(row=row_idx, column=5).value = concordancia
    ws3.cell(row=row_idx, column=6).value = conclusion
    
    # Estilos
    for col_idx in range(1, len(headers_comp) + 1):
        cell = ws3.cell(row=row_idx, column=col_idx)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx in [2, 3, 4]:
            cell.number_format = '0.0000E+00'

for col_idx in range(1, len(headers_comp) + 1):
    ws3.column_dimensions[get_column_letter(col_idx)].width = 18

# Guardar archivo
archivo_salida = '03_PASO3B_MANN_WHITNEY_U_N2480.xlsx'
wb.save(archivo_salida)
print(f"  ✓ Archivo guardado: {archivo_salida}\n")

# ════════════════════════════════════════════════════════════════════════════════
# PASO 5: RESUMEN EJECUTIVO
# ════════════════════════════════════════════════════════════════════════════════

print("════════════════════════════════════════════════════════════════════════════════")
print("RESUMEN EJECUTIVO - PASO 3B")
print("════════════════════════════════════════════════════════════════════════════════\n")

print("DATOS ANALIZADOS:")
print(f"  • Total de registros: {len(df):,}")
print(f"  • Manual: {len(manual_data):,} registros")
print(f"  • AI: {len(ia_data):,} registros")
print(f"  • Métricas: {len(metricas)}\n")

print("RESULTADOS MANN-WHITNEY U:")
for resultado in resultados_mw:
    print(f"\n  {resultado['Métrica']}:")
    print(f"    U-statistic: {resultado['U-statistic']:.2f}")
    print(f"    p-value: {resultado['p-value']:.6f}")
    print(f"    Significativo: {resultado['Significativo (α=0.05)']}")
    print(f"    Tamaño del efecto (r): {resultado['r (effect size)']:.4f}")

print("\n" + "="*80)
print("CONCLUSIÓN GENERAL:")
print("="*80)

# Contar significativos
significativos = sum(1 for r in resultados_mw if r['Significativo (α=0.05)'] == 'Sí')

print(f"\nDe {len(metricas)} métricas analizadas:")
print(f"  ✓ {significativos} son significativas (p < 0.05)")
print(f"  ✓ {len(metricas) - significativos} NO son significativas (p > 0.05)\n")

if significativos == 0:
    print("✅ CONCLUSIÓN: NO hay diferencias estadísticas significativas entre")
    print("   test cases Manual e IA en NINGUNA métrica (análisis con N=2,480).\n")
    print("   Esta conclusión es CONSISTENTE con los resultados de t-Student (N=12).\n")
    print("   → ANÁLISIS ROBUSTO Y VALIDADO ✓")
else:
    print("⚠️ ATENCIÓN: Se encontraron diferencias significativas.\n")
    print("   Revisar discrepancias con análisis de t-Student.")

print("\n" + "="*80)
print("✓ PASO 3B COMPLETADO")
print("="*80)
print(f"\nArchivos generados:")
print(f"  • {archivo_salida}")
print(f"\nPróximo paso: Revisar concordancia con t-Student en pestaña 'Comparacion_t-Student_vs_MW'\n")
